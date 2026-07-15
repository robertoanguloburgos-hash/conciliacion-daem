"""
Conciliador de Cartolas BCI
---------------------------
Aplicación Streamlit que rellena automáticamente la columna 'CARTOLA N°'
de un Excel de movimientos bancarios, cruzando cada movimiento contra los
montos extraídos de un PDF consolidado de cartolas oficiales del Banco BCI.

Todo el procesamiento ocurre en memoria (io.BytesIO), sin escritura a disco,
para ser 100% compatible con Streamlit Community Cloud.
"""

import streamlit as st
import pandas as pd
import pdfplumber
import openpyxl
import re
import io
from datetime import datetime

# =========================================================================
# CONFIGURACIÓN DE PÁGINA
# =========================================================================
st.set_page_config(
    page_title="Conciliador de Cartolas BCI",
    page_icon="🏦",
    layout="wide"
)

FILA_ENCABEZADO_DEFECTO = 7  # Fila 7 (1-indexada) = donde están los títulos reales de columnas

# =========================================================================
# EXPRESIONES REGULARES
# =========================================================================
PATRON_CARTOLA = re.compile(r"CARTOLA\s*N[°ºo]\s*(\d+)", re.IGNORECASE)
PATRON_FECHA = re.compile(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})")
# Montos con formato chileno: puntos de miles, coma opcional de decimales
PATRON_MONTO = re.compile(r"\d{1,3}(?:\.\d{3})+(?:,\d+)?|\d+,\d+")


# =========================================================================
# FUNCIONES AUXILIARES DE LIMPIEZA / PARSEO
# =========================================================================
def limpiar_monto(texto):
    """Convierte '33.592.180' o '33.592.180,50' en float. Devuelve None si no es numérico."""
    if texto is None:
        return None
    texto = str(texto).strip()
    if texto in ("", "-", "$"):
        return None
    texto = texto.replace("$", "").strip()
    texto = texto.replace(".", "")   # quita separador de miles
    texto = texto.replace(",", ".")  # coma -> punto decimal
    try:
        valor = float(texto)
        return valor
    except ValueError:
        return None


def parsear_fecha(texto):
    """Intenta parsear una fecha en los formatos comunes usados en cartolas BCI."""
    if texto is None:
        return None
    if isinstance(texto, datetime):
        return texto
    texto = str(texto).strip()
    formatos = ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"]
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


# =========================================================================
# EXTRACCIÓN DE MOVIMIENTOS DESDE EL PDF
# =========================================================================
def extraer_movimientos_pdf(pdf_bytes, progreso_callback=None):
    """
    Recorre el PDF página por página. Mantiene una variable de estado con la
    última 'CARTOLA N° X' detectada, de forma que todas las líneas leídas
    (incluso en páginas siguientes) se asignan a esa cartola hasta que se
    detecte una nueva.

    Devuelve un DataFrame con columnas: cartola, fecha, monto, texto_linea
    """
    registros = []
    cartola_actual = None

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total_paginas = len(pdf.pages)

        for i, pagina in enumerate(pdf.pages):
            texto = pagina.extract_text() or ""

            # Si en esta página aparece "CARTOLA N° X" una o más veces,
            # actualizamos el estado con la última ocurrencia de la página.
            coincidencias = PATRON_CARTOLA.findall(texto)
            if coincidencias:
                cartola_actual = int(coincidencias[-1])

            lineas = texto.split("\n")
            for linea in lineas:
                # No confundir la línea de encabezado de cartola con un movimiento
                if PATRON_CARTOLA.search(linea):
                    continue

                montos_encontrados = PATRON_MONTO.findall(linea)
                if not montos_encontrados or cartola_actual is None:
                    continue

                fechas_encontradas = PATRON_FECHA.findall(linea)
                fecha_linea = parsear_fecha(fechas_encontradas[0]) if fechas_encontradas else None

                for m in montos_encontrados:
                    valor = limpiar_monto(m)
                    if valor is not None and valor > 0:
                        registros.append({
                            "cartola": cartola_actual,
                            "fecha": fecha_linea,
                            "monto": valor,
                            "texto_linea": linea.strip()
                        })

            if progreso_callback:
                progreso_callback(i + 1, total_paginas)

    return pd.DataFrame(registros)


# =========================================================================
# LOCALIZACIÓN DE COLUMNAS EN EL EXCEL
# =========================================================================
def encontrar_columnas(ws, fila_encabezado):
    """Devuelve un dict {nombre_columna: indice_columna_1_indexado}."""
    columnas = {}
    for cell in ws[fila_encabezado]:
        if cell.value is not None:
            nombre = str(cell.value).strip()
            columnas[nombre] = cell.column
    return columnas


# =========================================================================
# ALGORITMO DE EMPAREJAMIENTO (FIFO + tolerancia de fecha)
# =========================================================================
def procesar_excel(excel_bytes, df_pdf, tolerancia_dias=3, fila_encabezado=FILA_ENCABEZADO_DEFECTO):
    """
    Carga el Excel con openpyxl (preserva formato/formulas), busca las
    columnas relevantes, y para cada fila de movimiento busca un match único
    en df_pdf por monto exacto + fecha dentro de tolerancia, usando FIFO y
    un set de índices ya usados para evitar duplicados.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    ws = wb.active

    columnas = encontrar_columnas(ws, fila_encabezado)

    col_fecha = columnas.get("Fecha contable (*)")
    col_cargo = columnas.get("Cargo (-)")
    col_abono = columnas.get("Abono (+)")
    col_cartola = columnas.get("CARTOLA N°")

    faltantes = [nombre for nombre, idx in [
        ("Fecha contable (*)", col_fecha),
        ("Cargo (-)", col_cargo),
        ("Abono (+)", col_abono),
        ("CARTOLA N°", col_cartola),
    ] if idx is None]

    if faltantes:
        raise ValueError(
            "No se encontraron las siguientes columnas esperadas en la fila "
            f"{fila_encabezado} del Excel: {faltantes}"
        )

    # Índice rápido: monto redondeado -> lista de índices en df_pdf
    pdf_por_monto = {}
    for idx, fila_pdf in df_pdf.iterrows():
        clave = round(fila_pdf["monto"], 2)
        pdf_por_monto.setdefault(clave, []).append(idx)

    used_pdf_indices = set()
    filas_emparejadas = 0
    filas_sin_match = 0

    total_filas = ws.max_row

    for fila in range(fila_encabezado + 1, total_filas + 1):
        cargo_val = ws.cell(row=fila, column=col_cargo).value
        abono_val = ws.cell(row=fila, column=col_abono).value
        fecha_val = ws.cell(row=fila, column=col_fecha).value

        monto_excel = None
        if cargo_val not in (None, "", 0):
            monto_excel = float(cargo_val) if isinstance(cargo_val, (int, float)) else limpiar_monto(cargo_val)
        elif abono_val not in (None, "", 0):
            monto_excel = float(abono_val) if isinstance(abono_val, (int, float)) else limpiar_monto(abono_val)

        if monto_excel is None or monto_excel == 0:
            continue

        fecha_excel = parsear_fecha(fecha_val)

        candidatos = pdf_por_monto.get(round(abs(monto_excel), 2), [])

        match_encontrado = None
        for idx_pdf in candidatos:
            if idx_pdf in used_pdf_indices:
                continue

            fecha_pdf = df_pdf.loc[idx_pdf, "fecha"]

            if fecha_excel is not None and fecha_pdf is not None:
                diferencia_dias = abs((fecha_excel - fecha_pdf).days)
                if diferencia_dias > tolerancia_dias:
                    continue

            # Primer candidato disponible (FIFO) que cumple la tolerancia
            match_encontrado = idx_pdf
            break

        if match_encontrado is not None:
            used_pdf_indices.add(match_encontrado)
            numero_cartola = df_pdf.loc[match_encontrado, "cartola"]
            ws.cell(row=fila, column=col_cartola).value = f"CARTOLA {numero_cartola}"
            filas_emparejadas += 1
        else:
            filas_sin_match += 1

    buffer_salida = io.BytesIO()
    wb.save(buffer_salida)
    buffer_salida.seek(0)

    return buffer_salida, filas_emparejadas, filas_sin_match


# =========================================================================
# INTERFAZ STREAMLIT
# =========================================================================
st.title("🏦 Conciliador de Cartolas BCI")
st.markdown(
    """
    Esta herramienta cruza automáticamente los movimientos del **Excel bancario**
    contra los montos extraídos del **PDF consolidado de Cartolas Oficiales BCI**,
    completando la columna **`CARTOLA N°`** según corresponda.
    """
)

st.divider()

col1, col2 = st.columns(2)
with col1:
    archivo_excel = st.file_uploader("📊 Excel de movimientos bancarios (.xlsx)", type=["xlsx"])
with col2:
    archivo_pdf = st.file_uploader("📄 PDF consolidado de Cartolas BCI (.pdf)", type=["pdf"])

with st.expander("⚙️ Opciones avanzadas"):
    fila_encabezado = st.number_input(
        "Fila donde están los encabezados reales del Excel",
        min_value=1, max_value=50, value=FILA_ENCABEZADO_DEFECTO,
        help="Por defecto es la fila 7, según la estructura estándar de la descarga BCI."
    )
    tolerancia = st.slider(
        "Tolerancia de días entre 'Fecha contable' del Excel y la fecha detectada en el PDF",
        min_value=0, max_value=10, value=3
    )

st.divider()

if archivo_excel and archivo_pdf:
    if st.button("🚀 Procesar y Conciliar", type="primary"):
        pdf_bytes = archivo_pdf.read()
        excel_bytes = archivo_excel.read()

        barra = st.progress(0, text="Iniciando extracción del PDF...")

        def actualizar_progreso(actual, total):
            barra.progress(actual / total, text=f"Leyendo página {actual} de {total} del PDF...")

        with st.spinner("Analizando el PDF de cartolas oficiales..."):
            df_pdf = extraer_movimientos_pdf(pdf_bytes, progreso_callback=actualizar_progreso)

        barra.empty()

        if df_pdf.empty:
            st.error(
                "No se pudo extraer ningún movimiento del PDF. Verifica que el archivo "
                "contenga el texto 'CARTOLA N° X' y montos con formato chileno (ej: 33.592.180)."
            )
        else:
            st.success(
                f"✅ Se extrajeron **{len(df_pdf)}** movimientos pertenecientes a "
                f"**{df_pdf['cartola'].nunique()}** cartolas distintas."
            )

            with st.spinner("Emparejando movimientos del Excel con el PDF..."):
                try:
                    buffer_resultado, emparejadas, sin_match = procesar_excel(
                        excel_bytes, df_pdf,
                        tolerancia_dias=tolerancia,
                        fila_encabezado=int(fila_encabezado)
                    )
                except ValueError as e:
                    st.error(f"❌ {e}")
                    st.stop()

            st.success(
                f"🎯 Proceso completado: **{emparejadas}** filas emparejadas, "
                f"**{sin_match}** filas sin coincidencia."
            )

            st.download_button(
                label="⬇️ Descargar Excel con CARTOLA N° completada",
                data=buffer_resultado,
                file_name="movimientos_bancarios_con_cartola.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            with st.expander("🔍 Ver detalle de movimientos extraídos del PDF"):
                st.dataframe(df_pdf, use_container_width=True)
else:
    st.info("⬆️ Sube ambos archivos (Excel y PDF) para habilitar el procesamiento.")
